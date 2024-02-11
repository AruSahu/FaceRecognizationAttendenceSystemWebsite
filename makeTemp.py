import openpyxl as op
import datetime
import takeAttendence
import sys
import openpyxl as op
import os
import csv


def createFile():
    filepath = "temp.xlsx"
    # print(filepath)
    wb = op.Workbook()
    wb.save(filepath)
    wb.close()





def defaultsTemp(time, subject, sem, batch):
    wb = op.load_workbook("temp.xlsx")
    wb["Sheet"].cell(row=1, column=1).value = "Time"
    wb["Sheet"].cell(row=1, column=2).value = time
    wb["Sheet"].cell(row=2, column=1).value = "Subject"
    wb["Sheet"].cell(row=2, column=2).value = subject
    wb["Sheet"].cell(row=1, column=4).value = "Sem"
    wb["Sheet"].cell(row=1, column=5).value = sem
    wb["Sheet"].cell(row=2, column=4).value = "Batch"
    wb["Sheet"].cell(row=2, column=5).value = batch
    wb["Sheet"].cell(row=1, column=7).value = "Year"
    year = datetime.datetime.today().strftime('%Y')
    month = datetime.datetime.today().strftime('%m')
    wb["Sheet"].cell(row=1, column=8).value = year
    wb["Sheet"].cell(row=2, column=7).value = "Date"
    str = datetime.datetime.today().strftime('%d-%m-%Y')
    print(str)
    wb["Sheet"].cell(row=2, column=8).value = str
    wb.save("temp.xlsx")
    month = int(month)
    # addStudentInTemp(studentPresnt)#list pass kar dena
    takeAttendence.mainAttendance(f"{batch}/{sem}_{year}{month}.xlsx", str)


def addStudentInTemp(studentPresent):
    file = "temp.xlsx"
    wbTemp = op.load_workbook(file)
    sheets = wbTemp.sheetnames
    j = 3
    for i in studentPresent:
        wbTemp[sheets[0]].cell(row=j, column=1).value = i
        j += 1
    wbTemp.save(file)


def readCSVFile():
    createFile()
    data = []

    with open('innovators.csv', mode='r') as file:
        # reading the CSV file
        csvFile = csv.reader(file)

        # displaying the contents of the CSV file
        for lines in csvFile:
            print(lines)
            data.append(lines[0])

    # data = data[1:]
    print(data)
    addStudentInTemp(data)


# studentPresentList = ["2019BTCS002","2019BTCS003","2019BTCS018", "2019BTCS049", "2019BTCS057", "2019BTCS069"] #2019BTCS

readCSVFile()

timinga = sys.argv[1]
subja = sys.argv[2]
sema = sys.argv[3]
yeara = sys.argv[4]

print(timinga)
print(subja)
print(sema)
print(yeara)
# dict - dict['year'], dict["semester"]
# defaultTEmp run karna h bas
defaultsTemp(timinga, subja, sema, yeara)  # Parameters from MySQL (time, subject, sem, batch, list)
# studentpresent = [18,57,49]
# makeTemp(studentpresent)
