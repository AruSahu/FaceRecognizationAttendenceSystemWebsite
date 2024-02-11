from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.models import User, auth
from .models import LoginINFO, AttendenceVideoFileUpload, AttendenceCSVFileUpload
from subprocess import run, PIPE

import datetime
import sys,requests, os
import openpyxl as op
import shutil

# Create your views here.

def index(request):

    return render(request, 'base/base.html')


def login(request):
    if request.method == 'POST':
        print("Entered wrongly 002")
        username = request.POST['username']
        password = request.POST['password']

        # user = auth.authenticate(username=username, password=password)
        loginInf = LoginINFO.objects.all()
        print("Important login info printing")
        tt1 = loginInf.values('loginID', 'password')
        print(tt1)
        print(tt1.values('password'))
        print(username)

        for i in loginInf.values('loginID', 'password', 'name', 'isAdmin'):  # call values needed for user login info
            print("Wrong Entry")
            if username == i['loginID'] and password == i['password']:
                # print("Partial success")
                name = i['name']
                # return render(request, 'loginApp/success.html', {'name': name})  # pass here to send to next page
                request.session['name'] = name
                if i['isAdmin'] == 1:
                    print("Reached 101")
                    return redirect('success_admin')
                elif i['isAdmin'] == 0:
                    print("Reached 102")
                    return redirect('success_teacher')


        # if user is not None:
        #     auth.login(request, user)
        #     return redirect('success')
        # passed login
        else:
            return redirect('failed')
    else:
        print("Entered wrongly 001")
        return render(request, 'loginApp/logint.html')


def success_admin(request):
    return render(request, 'loginApp/success_admin.html')


def success_teacher(request):
    return render(request, 'loginApp/success_teacher.html')


def failed(request):
    return render(request, 'loginApp/failed.html')


def logout(request):
    print("Issue 001")
    request.session['name'] = 'NULL'
    # return render(request, 'loginApp/login.html')
    return redirect('login')

def takeAttendence(request):
    print("Issue 1101")
    request.session['name']
    # return redirect('takeAttendence')

    return render(request, 'loginApp/takeAttendence.html')

def send_files_csv(request):
    if request.method == "POST":
        teacherName = request.POST.get("nameTeacher")
        sem_csv = request.POST.get("semester_csv")
        year_csv = request.POST.get("batch_csv")
        subject_csv = request.POST.get("subject_csv")
        timing_csv = request.POST.get("timing_csv")
        csvfile = request.FILES.getlist("filecsv")
        for f in csvfile:
            AttendenceCSVFileUpload(nameTeacher=teacherName, semester_csv=sem_csv, year_csv=year_csv, subject_csv=subject_csv, timing_csv=timing_csv, csvfile=f).save()

        dll_csv = AttendenceCSVFileUpload.objects.all()
        ddf_csv = dll_csv.values('csvfile')
        ddp_csv = ddf_csv[ddf_csv.count() - 1]['csvfile']
        print(ddp_csv)

        shutil.copyfile('video/'+ddp_csv, 'innovators.csv')

        out = run([sys.executable, "makeTemp.py", timing_csv, subject_csv, sem_csv, year_csv], shell=False, stdout=PIPE)
        print(out)

        os.remove("temp.xlsx")
        os.remove("innovators.csv")


    return redirect(success_page)

def send_files(request):
    if request.method == "POST":
        teacherName = request.POST.get("nameTeacher")
        sem = request.POST.get("semester")
        year = request.POST.get("year")
        subject = request.POST.get("subject")
        timing = request.POST.get("timing")
        videofile = request.FILES.getlist("filevideo")
        for f in videofile:
            AttendenceVideoFileUpload(nameTeacher=teacherName, semester=sem, year=year, subject=subject, timing=timing, videofile=f).save()
        print("find here")
        print(f)
        dll = AttendenceVideoFileUpload.objects.all()
        ddf = dll.values('videofile')
        ddp = ddf[ddf.count()-1]['videofile']
        print(ddp)

        # return HttpResponse("Video Uploaded to the Server Successfully, please wait while the attendence is being recorded")
        out =run([sys.executable, "face_recognition.py", ddp], shell=False, stdout=PIPE)
        # print(out.decode("utf-8"))
        print(out)

        al = AttendenceVideoFileUpload.objects.all()
        tmrp = al.values('timing', 'subject', 'semester', 'year')
        ttime = tmrp[tmrp.count()-1]['timing']
        tsubj = tmrp[tmrp.count()-1]['subject']
        tsem = tmrp[tmrp.count()-1]['semester']
        tyear = tmrp[tmrp.count()-1]['year']
        print(tsubj)
        print(ttime)
        print(tsem)
        print(tyear)
        out = run([sys.executable, "makeTemp.py", ttime, tsubj, tsem, tyear], shell=False, stdout=PIPE)
        print(out)

        os.remove("temp.xlsx")
        os.remove("innovators.csv")
        return redirect(success_page)
        # return render(request, "loginApp/h.html", {'data1': out.stdout.decode('utf-8')})

def success_page(request):
    return render(request, 'loginApp/success.html')

def checkAttendenceForm(request):
    # request.session['name'] = 'NULL'
    return render(request, 'loginApp/checkAttendenceForm.html')

def checkAttendence(request):
    # request.session['name'] = 'NULL'
    if request.method == "POST":
        teacherName = request.POST.get("nameTeacher")
        sem = request.POST.get("semester")
        batch = request.POST.get("batch")
        subject = request.POST.get("subject")
        timing = request.POST.get("timing")
        dateb = request.POST.get("date")

        datea = dateb

        dateb = dateb.split("-")
        dateb = dateb[-1::-1]
        dateb = "-".join(dateb)

        datem = datetime.datetime.strptime(datea, "%Y-%m-%d")
        datemonth = datem.month
        dateyear = datem.year

        print("My date format")
        # print(datea)
        print(dateb)


        file = f"{batch}/{sem}_{dateyear}{datemonth}.xlsx"
        wb = op.load_workbook(file)
        dateColumn = 6
        myDate = wb[subject].cell(row=2, column=dateColumn).value
        # myDate = myDate.split("-")
        # myDate = myDate[-1::-1]
        # myDate = "-".join(myDate)
        print(myDate)
        while myDate != dateb and myDate!=None:
            # print(wb[sub].cell(row=2, column=dateColumn).value)
            dateColumn += 1
            myDate = wb[subject].cell(row=2, column=dateColumn).value
            # myDate = myDate.split("-")
            # myDate = myDate[-1::-1]
            # myDate = "-".join(myDate)
            print(myDate)
        datah = []
        j = 3
        while wb[subject].cell(row=j, column=2).value is not None:
            name = str(wb[subject].cell(row=j, column=1).value)
            enNO = str(wb[subject].cell(row=j, column=2).value)
            pa = str(wb[subject].cell(row=j, column=dateColumn).value)
            if pa == 'None':
                pa = ""
            temp = {'id': j - 2, 'enroll': enNO, 'name': name, 'pa': pa}
            datah.append(temp)
            print(temp)
            j += 1



    return render(request, 'loginApp/checkAttendance.html', {'teacherName': teacherName,'sem': sem, 'batch': batch, 'subject': subject, 'timing': timing, 'datea': dateb, 'datemonth': datemonth, 'datah': datah})
