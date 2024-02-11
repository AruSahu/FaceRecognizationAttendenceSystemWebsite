import openpyxl as op
import os
import time


def allStudentNames():
    studentNames = op.load_workbook('2019/Students.xlsx')
    sheetNames = studentNames.active

    for i in range(94):
        enNo = sheetNames.cell(row=i + 2, column=1).value
        name = sheetNames.cell(row=i + 2, column=2).value
        print(enNo, name)


def makeSheets(filepath, subjects):
    os.rename(filepath, filepath.split('/')[1])
    mainFilePath = filepath
    filepath = filepath.split('/')[1]
    wb = op.Workbook(filepath)
    for i in range(len(subjects)):
        new_sheet = wb.create_sheet(subjects[i], i)
    wb.save(filepath)
    wb.close()
    os.rename(filepath, mainFilePath)


def createFile(sem, batch, year, month):
    filepath = f"{sem}_{year}{month}.xlsx"
    if not os.path.isfile(filepath):
        # print(filepath)
        wb = op.Workbook()
        wb.save(filepath)
        wb.close()
        os.rename(filepath, f"{batch}/{filepath}")


def setDefaults(file, batch):
    wb = op.load_workbook(file)
    sheets = wb.sheetnames
    print(sheets)
    print("Set Defaults")
    studentNames = op.load_workbook(f'{batch}/Students.xlsx')
    sheetNames = studentNames.active
    enNo = []
    name = []
    for i in range(94):
        enNo.append(sheetNames.cell(row=i + 2, column=1).value)
        name.append(sheetNames.cell(row=i + 2, column=2).value)
        # print(enNo, name)

    # print(wb["TOC"].cell(row=1, column=1).value)
    for i in sheets:
        wb[i].cell(row=2, column=1).value = "Name"
        wb[i].cell(row=2, column=2).value = "Enrollment Number"
        wb[i].cell(row=2, column=3).value = "Total Classes"
        wb[i].cell(row=2, column=4).value = "Present Class"
        wb[i].cell(row=2, column=5).value = "Percentage"
        for j in range(94):
            wb[i].cell(row=3 + j, column=1).value = name[j]
            wb[i].cell(row=3 + j, column=2).value = enNo[j]
            wb[i].cell(row=3 + j, column=3).value = "=COUNTA(F2:Z2)"
            wb[i].cell(row=3 + j, column=4).value = f"=COUNTIF(F{j + 3}:Z{j + 3}, \"P\")"
            wb[i].cell(row=3 + j, column=5).value = f"=IF(C{j + 3} = 0, 0, (D{j + 3}/C{j + 3})*100)"
    wb.save(file)
    wb.close()


def makeTheMainFile():
    x = input("Enter the sem, batch, year and the month").split()
    sem, batch, year, month = int(x[0]), int(x[1]), int(x[2]), int(x[3])  # Set the sem, year, month

    createFile(sem, batch, year, month)  # Create the file

    filepath = f"{sem}_{year}{month}.xlsx"
    subjects = input("Enter the Subjects: ").split()

    makeSheets(f"{batch}/{filepath}", subjects)  # Create the diffrent sheets
    # time.sleep(5)
    setDefaults(f"{batch}/{filepath}", batch)

# RUN THIS FILE MONTHLY (We can make this automatic)
makeTheMainFile()