import openpyxl as op


def setDefaultsFunc(batch, sem, year, month):
    file = f"{batch}/{sem}_{year}{month}.xlsx"
    wb = op.load_workbook(file)
    sheets = wb.sheetnames
    print(sheets)
    print("Set Defaults")
    studentNames = op.load_workbook('2019/Students.xlsx')
    sheetNames = studentNames.active
    enNo = []
    name = []
    for i in range(94):
        enNo.append(sheetNames.cell(row=i + 2, column=1).value)
        name.append(sheetNames.cell(row=i + 2, column=2).value)
        # print(enNo, name)

    # print(wb["TOC"].cell(row=1, column=1).value)
    for i in sheets:
        wb[i].cell(row=2, column=1).value = "Name"
        wb[i].cell(row=2, column=2).value = "Enrollment Number"
        wb[i].cell(row=2, column=3).value = "Total Classes"
        wb[i].cell(row=2, column=4).value = "Present Class"
        wb[i].cell(row=2, column=5).value = "Percentage"
        for j in range(94):
            wb[i].cell(row=3 + j, column=1).value = name[j]
            wb[i].cell(row=3 + j, column=2).value = enNo[j]
            wb[i].cell(row=3 + j, column=3).value = "=COUNTA(F2:Z2)"
            wb[i].cell(row=3 + j, column=4).value = f"=COUNTIF(F{j + 3}:Z{j + 3}, \"P\")"
            wb[i].cell(row=3 + j, column=5).value = f"=IF(C{j + 3} = 0, 0, (D{j + 3}/C{j + 3})*100)"
    wb.save(file)


batch = 2019
sem = 8
year = 2022
month = 1
setDefaultsFunc(batch, sem, year, month)
