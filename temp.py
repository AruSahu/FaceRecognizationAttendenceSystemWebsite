import openpyxl as op
import os

def createFile():
    filepath = "temp.xlsx"
    # print(filepath)
    wb = op.Workbook()
    wb.save(filepath)
    wb.close()

def defaultsTemp(time, subject, sem, batch, year, date):
    wb = op.load_workbook("temp.xlsx")
    wb["Sheet"].cell(row=1, column=1).value = "Time"
    wb["Sheet"].cell(row=1, column=2).value = time
    wb["Sheet"].cell(row=2, column=1).value = "Subject"
    wb["Sheet"].cell(row=2, column=2).value = subject
    wb["Sheet"].cell(row=1, column=4).value = "Sem"
    wb["Sheet"].cell(row=1, column=5).value = sem
    wb["Sheet"].cell(row=2, column=4).value = "Batch"
    wb["Sheet"].cell(row=2, column=5).value = batch
    wb["Sheet"].cell(row=1, column=7).value = "Year"
    wb["Sheet"].cell(row=1, column=8).value = year
    wb["Sheet"].cell(row=2, column=7).value = "Date"
    wb["Sheet"].cell(row=2, column=8).value = date
    wb.save("temp.xlsx")

defaultsTemp("09:30-11:30", "IP", "V", "2019", "2022", "18-01-2022")