import openpyxl as op

def mainAttendance(file, datecheck):
    wb = op.load_workbook(file)

    temp = "temp.xlsx"
    wTemp = op.load_workbook(temp)

    subjectAttendance = wTemp["Sheet"].cell(row=2, column=2).value
    dateAttendance = wTemp["Sheet"].cell(row=2, column=8).value

    print(subjectAttendance, dateAttendance)

    # new date add karne ke liye
    dateColumn = 6
    while wb[subjectAttendance].cell(row=2, column=dateColumn).value is not None and wb[subjectAttendance].cell(row=2, column=dateColumn).value != datecheck:
        print(wb[subjectAttendance].cell(row=2, column=dateColumn).value)
        dateColumn += 1

    wb[subjectAttendance].cell(row=2, column=dateColumn).value = str(dateAttendance).split()[0]
    # print(wb[subjectAttendance].cell(row=2, column=dateColumn).value)

    # Making the list of presenties
    j = 3
    presentStudents = []
    while wTemp["Sheet"].cell(row=j, column=1).value is not None:
        # presentStudent = wTemp["Sheet1"].cell(row=j, column=1).value
        # print(presentStudent)
        # presentStudents.append(presentStudent)
        presentStudents.append(wTemp["Sheet"].cell(row=j, column=1).value)
        j += 1

    # print(presentStudents.sort())
    j = 3
    while wb[subjectAttendance].cell(row=j, column=2).value is not None:
        # print(j)
        # j += 1
        enNO = str(wb[subjectAttendance].cell(row=j, column=2).value)
        # print("enNO ", enNO)
        i = enNO
        # print(j)
        # print(i)
        if i in presentStudents:
            # print(i)
            wb[subjectAttendance].cell(row=j, column=dateColumn).value = "P"
        j += 1
    # # print(i)
    wb.save(file)


# mainAttendance("2019/5_20221.xlsx")