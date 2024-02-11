import openpyxl as op

studentPresentList = ["2019BTCS002","2019BTCS003","2019BTCS018", "2019BTCS049", "2019BTCS057", "2019BTCS069"]

def createFile():
    filepath = "tempr.xlsx"
    # print(filepath)
    wb = op.Workbook()
    wb.save(filepath)
    wb.close()

def addStudentInTemp(studentPresent):
    createFile()
    file = "tempr.xlsx"
    wbTemp = op.load_workbook(file)
    j = 3
    for i in studentPresent:
        wbTemp["Sheet"].cell(row=j, column=1).value = i
        j += 1
    wbTemp.save(file)

addStudentInTemp(studentPresentList)